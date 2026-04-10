import os
import traceback
import io
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Optional, List, Union, Any
from typing_extensions import override

# --- DOCX IMPORTS ---
from docx import Document as DocxDocument
from docx.document import Document as _Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table, _Cell
from docx.text.paragraph import Paragraph

from mcp.server.fastmcp import FastMCP, Image
from openpyxl import load_workbook
import fitz  # PyMuPDF
from PIL import Image as PILImage

# Allow processing of huge images (like 200 Mpix)
PILImage.MAX_IMAGE_PIXELS = None

mcp = FastMCP("Document Reader")

# Configuration for image processing
# We limit by total megapixels to balance resolution and transfer size
MAX_TOTAL_PIXELS = 6_000_000  # ~6 Megapixels (e.g., 2449x2449 or equivalent area)
TARGET_DPI = 300              # DPI for rendering PDF pages

def process_image_for_transport(image_bytes: bytes) -> bytes:
    """
    Resizes and compresses images to ensure they fit within MCP transport limits
    while preserving legibility of large drawings.
    """
    try:
        img = PILImage.open(io.BytesIO(image_bytes))

        # Calculate scaling factor based on area to preserve extreme aspect ratios
        current_pixels = img.width * img.height
        if current_pixels > MAX_TOTAL_PIXELS:
            scale = (MAX_TOTAL_PIXELS / current_pixels) ** 0.5
            new_width = int(img.width * scale)
            new_height = int(img.height * scale)
            img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)

        # Convert to RGB (removes alpha channel which saves space)
        if img.mode != "RGB":
            img = img.convert("RGB")

        output = io.BytesIO()
        # Save as JPEG with high quality to keep text readable but file size small
        img.save(output, format="JPEG", quality=85, optimize=True)
        return output.getvalue()
    except Exception as e:
        print(f"Image processing error: {e}")
        return image_bytes

class DocumentReader(ABC):
    @abstractmethod
    def read(self, file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
        pass

class DocxReader(DocumentReader):
    """Handles DOCX structure and images."""

    def _get_block_elements(self, parent):
        if isinstance(parent, _Document):
            parent_elm = parent.element.body
        elif isinstance(parent, _Cell):
            parent_elm = parent._tc
        else:
            return
        for child in parent_elm.iterchildren():
            if isinstance(child, CT_P): yield Paragraph(child, parent)
            elif isinstance(child, CT_Tbl): yield Table(child, parent)

    def _process_content(self, container, level=0) -> List[str]:
        output = []
        indent = "  " * level
        for block in self._get_block_elements(container):
            if isinstance(block, Paragraph):
                if block.text.strip(): output.append(f"{indent}{block.text.strip()}")
            elif isinstance(block, Table):
                output.append(f"{indent}[TABLE]")
                for row in block.rows:
                    row_data = [ " ".join(self._process_content(cell, level + 1)).strip() for cell in row.cells ]
                    output.append(f"{indent}  | {' | '.join(row_data)} |")
        return output

    @override
    def read(self, file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
        try:
            doc = DocxDocument(file_path)
            all_text = self._process_content(doc)
            results = ["\n".join(all_text[:100])] # Simplified text return

            # Extract images from DOCX
            img_count = 0
            for rel in doc.part.rels.values():
                if "image" in rel.target_ref and img_count < 3:
                    processed = process_image_for_transport(rel.target_part.blob)
                    results.append(Image(data=processed, format="image/jpeg"))
                    img_count += 1
            return results
        except Exception as e:
            return [f"Docx Error: {str(e)}"]

class PdfReader(DocumentReader):
    """
    Renders PDF pages as images to handle large drawings properly
    instead of extracting raw (often tiled) image objects.
    """
    @override
    def read(self, file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
        try:
            doc = fitz.open(file_path)
            total_pages = len(doc)
            start_idx = max(0, start_page - 1)
            end_idx = min(end_page if end_page else start_page, total_pages)

            results: List[Union[str, Image]] = []
            results.append(f"--- PDF: {total_pages} pages. Showing {start_idx+1}-{end_idx} ---")

            for page_num in range(start_idx, end_idx):
                page = doc[page_num]

                # 1. Get Text
                text = page.get_text("text")
                results.append(f"\n--- PAGE {page_num + 1} ---\n{text}")

                # 2. Render Full Page as Image (Best for drawings/maps)
                # We use a Matrix to scale to the desired DPI
                zoom = TARGET_DPI / 72
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)

                # Convert Pixmap to bytes
                img_bytes = pix.tobytes("jpeg")

                # Further scale and compress to fit transport limits
                processed_img = process_image_for_transport(img_bytes)
                results.append(Image(data=processed_img, format="image/jpeg"))

            doc.close()
            return results
        except Exception as e:
            return [f"PDF Error: {str(e)}"]

class TxtReader(DocumentReader):
    @override
    def read(self, file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            return [content[:20000]] # Limit return size
        except Exception as e:
            return [f"Txt Error: {str(e)}"]

class ExcelReader(DocumentReader):
    @override
    def read(self, file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
        try:
            wb = load_workbook(file_path, read_only=True)
            res = []
            for sheet in wb.worksheets:
                res.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(max_row=50, values_only=True):
                    res.append("\t".join([str(c) if c else "" for c in row]))
            return ["\n".join(res)[:15000]]
        except Exception as e:
            return [f"Excel Error: {str(e)}"]

class DocumentReaderFactory:
    _readers = {
        ".txt": TxtReader, ".tex": TxtReader, ".docx": DocxReader,
        ".pdf": PdfReader, ".xlsx": ExcelReader, ".xls": ExcelReader,
    }

    @classmethod
    def get_reader(cls, file_path: str) -> DocumentReader:
        ext = os.path.splitext(file_path.lower())[1]
        if ext not in cls._readers: raise ValueError(f"Unsupported: {ext}")
        return cls._readers[ext]()

@mcp.tool()
def read_document(filename: str, start_page: int = 1, end_page: Optional[int] = None) -> Any:
    """Reads documents. PDF pages are rendered as images to handle large drawings/maps."""
    path = Path(filename)
    if not path.exists(): return [f"Error: {filename} not found."]
    try:
        reader = DocumentReaderFactory.get_reader(str(path))
        return reader.read(str(path), start_page=start_page, end_page=end_page)
    except Exception as e:
        return [f"Error: {str(e)}"]

if __name__ == "__main__":
    mcp.run()